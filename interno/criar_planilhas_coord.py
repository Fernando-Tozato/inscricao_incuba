import os

from django.db.models import QuerySet
from django.db.models.functions import Coalesce
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from database.models import Aluno, Turma


def ajustar_colunas(ws: Worksheet):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Obtém a letra da coluna
        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width


def criar_folha_geral(wb: Workbook):
    ws_geral = wb.create_sheet(title='Geral')
    headers = ['ID', 'Nome', 'CPF', 'Celular', 'Rua', 'Número', 'Bairro', 'Cidade', 'UF', 'É PCD?', 'OBS', 'Curso', 'Dias', 'Horario']
    for col_num, header in enumerate(headers, start=1):
        ws_geral.cell(row=1, column=col_num).value = header
        ws_geral.cell(row=1, column=col_num).font = Font(bold=True)
        ws_geral.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")

    alunos = Aluno.objects.all()

    for row_num, aluno in enumerate(alunos, start=2):
        turma = aluno.id_turma

        ws_geral.cell(row=row_num, column=1).value = aluno.pk
        ws_geral.cell(row=row_num, column=2).value = aluno.nome_social if aluno.nome_social else aluno.nome
        ws_geral.cell(row=row_num, column=3).value = aluno.cpf_formatado()
        ws_geral.cell(row=row_num, column=4).value = aluno.celular
        ws_geral.cell(row=row_num, column=5).value = aluno.rua
        ws_geral.cell(row=row_num, column=6).value = aluno.numero
        ws_geral.cell(row=row_num, column=7).value = aluno.bairro
        ws_geral.cell(row=row_num, column=8).value = aluno.cidade
        ws_geral.cell(row=row_num, column=9).value = aluno.uf
        ws_geral.cell(row=row_num, column=10).value = 'PCD' if aluno.pcd else ''
        ws_geral.cell(row=row_num, column=11).value = aluno.observacoes if aluno.observacoes is not None else ''
        ws_geral.cell(row=row_num, column=12).value = turma.curso
        ws_geral.cell(row=row_num, column=13).value = turma.dias
        ws_geral.cell(row=row_num, column=14).value = turma.horario()

        for col in range(1, 15):
            ws_geral.cell(row=row_num, column=col).alignment = Alignment(horizontal="left")

    ajustar_colunas(ws_geral)
    return ws_geral


def criar_folha_curso(wb: Workbook, curso: str, turmas: QuerySet[Turma]):
    dark_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")

    ws_curso = wb.create_sheet(title=f'{curso}')
    headers = ['ID', 'Nome', 'CPF', 'Celular', 'Rua', 'Número', 'Bairro', 'Cidade', 'UF', 'É PCD?', 'OBS']

    init_row = 1
    for turma in turmas:
        last_row = init_row

        # Separação de turmas
        ws_curso.cell(row=init_row, column=1).fill = dark_fill
        ws_curso.cell(row=init_row, column=2).value = f'{turma.dias} - {turma.horario()}'
        for col in range(3, 12):
            ws_curso.cell(row=init_row, column=col).fill = dark_fill

        # Headers
        for col_num, header in enumerate(headers, start=1):
            ws_curso.cell(row=init_row + 1, column=col_num).value = header
            ws_curso.cell(row=init_row + 1, column=col_num).font = Font(bold=True)
            ws_curso.cell(row=init_row + 1, column=col_num).alignment = Alignment(horizontal="center")

        alunos = (Aluno.objects.filter(id_turma=turma)
                  .annotate(nome_ordenacao=Coalesce('nome_social', 'nome'))
                  .order_by('nome_ordenacao'))

        last_row += 2
        for row_num, aluno in enumerate(alunos, start=last_row):
            ws_curso.cell(row=row_num, column=1).value = aluno.pk
            ws_curso.cell(row=row_num, column=2).value = aluno.nome_social if aluno.nome_social else aluno.nome
            ws_curso.cell(row=row_num, column=3).value = aluno.cpf_formatado()
            ws_curso.cell(row=row_num, column=4).value = aluno.celular
            ws_curso.cell(row=row_num, column=5).value = aluno.rua
            ws_curso.cell(row=row_num, column=6).value = aluno.numero
            ws_curso.cell(row=row_num, column=7).value = aluno.bairro
            ws_curso.cell(row=row_num, column=8).value = aluno.cidade
            ws_curso.cell(row=row_num, column=9).value = aluno.uf
            ws_curso.cell(row=row_num, column=10).value = 'PCD' if aluno.pcd else ''
            ws_curso.cell(row=row_num, column=11).value = aluno.observacoes if aluno.observacoes is not None else ''

            for col in range(1, 12):
                ws_curso.cell(row=row_num, column=col).alignment = Alignment(horizontal="left")
            last_row += 1

        for col_num in range(1, 12):
            ws_curso.cell(row=last_row, column=col_num).fill = dark_fill

        init_row = last_row + 1

    print(ws_curso)
    ajustar_colunas(ws_curso)
    return ws_curso


def gerar_planilha_coord():
    curso_title = {
        'Informática para Iniciantes e Melhor Idade': 'Info Iniciante',
        'Informática Básica': 'Info Básica',
        'Excel': 'Excel',
        'Montagem e Manutenção de Computadores': 'Montagem',
        'Robótica e Automação: Módulo 01': 'Robótica 1',
        'Robótica e Automação: Módulo 02': 'Robótica 2',
        'Robótica e Automação: Módulo 03': 'Robótica 3',
        'Design e Modelagem 3D: Módulo 01': 'Design 1',
        'Design e Modelagem 3D: Módulo 02': 'Design 2',
        'Gestão de Pessoas': 'Gestão',
        'Educação Financeira': 'Educação Financeira',
        'Marketing Digital': 'M. Digital',
        'Marketing Empreendedor': 'M. Empreendedor'
    }

    try:
        wb = Workbook()

        default_sheet = wb.active
        wb.remove(default_sheet)

        criar_folha_geral(wb)

        turmas = Turma.objects.all()
        cursos = set(turma.curso for turma in turmas)

        for curso in cursos:
            turmas_aux = turmas.filter(curso=curso)

            if not turmas_aux.exists():
                continue

            criar_folha_curso(wb, curso_title[curso], turmas_aux)

        nome_arquivo = f'planilha_geral_2024_2.xlsx'
        caminho_arquivo = os.path.join('media', 'planilha_coord', nome_arquivo)

        os.makedirs(os.path.dirname(caminho_arquivo), exist_ok=True)

        wb.save(caminho_arquivo)
        print(f'Planilha salva: {caminho_arquivo}')

    except Exception as e:
        print(f'Ocorreu um erro durante a geração da planilha: {e}')
