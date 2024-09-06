from django.db.models import Q, QuerySet
from django.db.models.functions import Coalesce
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet

from database.models import Aluno, Turma


class Alunos:
    ws: Worksheet
    professor: str
    curso: str

    def __init__(self, ws: Worksheet, professor: str, curso: str) -> None:
        self.ws = ws
        self.professor = professor
        self.curso = curso
        self.generate()

    def generate(self) -> Worksheet:
        turmas: QuerySet[Turma] = Turma.objects.filter(Q(professor=self.professor) & Q(curso=self.curso))
        alunos: QuerySet[Aluno] = (Aluno.objects.filter(id_turma__in=turmas)
                                   .annotate(nome_ordenacao=Coalesce('nome_social', 'nome'))
                                   .order_by('nome_ordenacao'))

        self.criar_headers()

        for row, aluno in enumerate(alunos, start=2):
            self.ws.cell(row=row, column=1).value = aluno.pk
            self.ws.cell(row=row, column=2).value = aluno.nome_social if aluno.nome_social else aluno.nome
            self.ws.cell(row=row, column=3).value = aluno.cpf
            self.ws.cell(row=row, column=4).value = aluno.celular
            self.ws.cell(row=row, column=5).value = aluno.rua
            self.ws.cell(row=row, column=6).value = aluno.numero
            self.ws.cell(row=row, column=7).value = aluno.bairro
            self.ws.cell(row=row, column=8).value = aluno.cidade
            self.ws.cell(row=row, column=9).value = aluno.uf
            self.ws.cell(row=row, column=10).value = 'PCD' if aluno.pcd else ''

        return self.ws

    def criar_headers(self) -> None:
        headers: list = ['ID', 'Nome', 'CPF', 'Celular', 'Rua', 'Número', 'Bairro', 'Cidade', 'UF', 'É PCD?']

        for col, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
