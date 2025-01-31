import json
import os
import random
import zipfile
from datetime import timedelta
from logging import getLogger, Logger

from celery import shared_task, group
from django.conf import settings
from django.core.mail import EmailMessage
from django.db.models import Q, QuerySet
from django.db.models.functions import Coalesce
from django.template.loader import render_to_string
from django_celery_beat.models import IntervalSchedule, PeriodicTask
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from database.models import *
from .middleware import get_current_user

'''
    ENVIO DE EMAILS
'''
# Task para enviar lista de emails com conteúdo e assunto, HTML e anexo opcionais
@shared_task
def enviar_emails(**kwargs):
    # Parâmetros obrigatórios
    emails: list[str]       = kwargs.get('emails')
    content: str            = kwargs.get('content')
    subject: str            = kwargs.get('subject')

    # Parâmetros opcionais
    has_html: bool          = kwargs.get('has_html', False)
    file_path: [str, None]  = kwargs.get('file_path', None)

    for email in emails:
        email_message: EmailMessage = EmailMessage(
            subject,
            content,
            settings.EMAIL_HOST_USER,
            [email],
        )

        if has_html:
            email_message.content_subtype = 'html'

        if file_path is not None:
            email_message.attach_file(file_path)

        email_message.send()


'''
    SORTEIO E AVISO DE INSCRITOS
'''

@shared_task
def sortear():
    logger = getLogger('sorteio')
    controle = Controle.objects.first()
    turmas = Turma.objects.all()
    inscritos = Inscrito.objects.all()

    sorteio(logger, turmas, inscritos)

    emails_sorteados = list(inscritos.filter(Q(ja_sorteado=True) & Q(email__isnull=False)).values_list('email', flat=True))
    emails_nao_sorteados = list(inscritos.filter(Q(ja_sorteado=False) & Q(email__isnull=False)).values_list('email', flat=True))

    content_sorteados = render_to_string("emails/aviso_sorteado.html", {
        'domain': 'incubarobotica.com.br',
        'protocol': 'https',
        'matricula_sorte_inicio': controle.matricula_sorte_inicio,  # type: ignore
        'matricula_sorte_fim': controle.matricula_sorte_fim,  # type: ignore
        'matricula_reman_inicio': controle.matricula_reman_inicio,  # type: ignore
        'matricula_reman_fim': controle.matricula_reman_fim,  # type: ignore
        'aulas_inicio': controle.aulas_inicio,  # type: ignore
    })

    content_nao_sorteados = render_to_string("emails/aviso_nao_sorteado.html", {
        'domain': 'incubarobotica.com.br',
        'protocol': 'https',
        'matricula_reman_inicio': controle.matricula_reman_inicio,  # type: ignore
        'matricula_reman_fim': controle.matricula_reman_fim,  # type: ignore
        'aulas_inicio': controle.aulas_inicio,  # type: ignore
    })

    tarefas = [
        preparar_emails_sorteio.s(emails_sorteados, content_sorteados),
        preparar_emails_sorteio.s(emails_nao_sorteados, content_nao_sorteados)
    ]
    group(tarefas).apply_async()

def sorteio(logger: Logger, turmas: QuerySet[Turma], inscritos: QuerySet[Inscrito]):
    """
    Realiza o sorteio de inscritos em turmas conforme as regras especificadas.

    Este código é público para fins de transparência e validação do sorteio.

    - Para cada turma, o sorteio é realizado separadamente para vagas de cotas e de ampla concorrência.
    - Os inscritos sorteados são marcados como 'ja_sorteado' para evitar múltiplas seleções.
    - As operações são registradas em 'sorteio.log' para auditoria e verificação posterior.
    """

    logger.info("Inicio do sorteio")  # Mostra o início do sorteio no log

    for turma in turmas:  # Itera sobre todas as turmas
        logger.info(f"Turma:{turma.unidade} | {turma.curso.nome} | {turma.dias} | {turma.horario()}")  # Mostra no log a turma atual

        """
        Sorteio para vagas de cotas
        """

        vagas_cotas = turma.cotas()  # Pega o número de vagas para cotas (30% do total das vagas)
        inscritos_cotas = list(inscritos.filter(Q(id_turma=turma) & Q(Q(pcd=True) | Q(
            ps=True))))  # Pega todos os inscritos para a turma atual que sejam PCD ou participem de Programa Social

        if len(inscritos_cotas) <= vagas_cotas:  # Caso o número de inscritos cotistas para a turma atual seja menor
            # que o número de vagas para cotas da turma...
            sorteados_cotas = inscritos_cotas  # ... define os sorteados cotistas para a turma atual como todos os
            # inscritos para a turma atual
        else:  # Se não...
            sorteados_cotas = random.sample(inscritos_cotas,
                                            vagas_cotas)  # ... sorteia os inscritos de acordo com o número de vagas

        for sorteado in sorteados_cotas:  # Itera sobre todos os sorteados cotistas da turma atual
            logger.info(
                f'{sorteado.nome_social if sorteado.nome_social else sorteado.nome} | {sorteado.num_inscricao_formatado()} | Cota')  # Mostra o nome civil ou social do sorteado no log
            sorteado.ja_sorteado = True  # Define o atributo ja_sorteado como verdadeiro
            sorteado.save()  # Salva a alteração

        """
        Sorteio para vagas de ampla concorrência
        """

        vagas_gerais = turma.ampla_conc()  # Pega o número de vagas para ampla concorrência (70% do total das vagas)
        inscritos_gerais = list(inscritos.filter(Q(id_turma=turma) & Q(
            ja_sorteado=False)))  # Pega todos os inscritos para a turma atual que não tenham sido sorteados ainda

        if len(inscritos_gerais) <= vagas_gerais:  # Caso o número de inscritos para a turma atual seja menor que o
            # número de vagas da turma...
            sorteados_gerais = inscritos_gerais  # ... define os sorteados para a turma atual como todos os inscritos
            # para a turma atual
        else:  # Se não...
            sorteados_gerais = random.sample(inscritos_gerais,
                                             vagas_gerais)  # ... sorteia os inscritos de acordo com o número de vagas

        for sorteado in sorteados_gerais:  # Itera sobre todos os sorteados de ampla concorrência da turma atual
            logger.info(
                f'{sorteado.nome_social if sorteado.nome_social else sorteado.nome} | {sorteado.num_inscricao_formatado()} | Ampla concorrência')  # Mostra o nome civil ou social do sorteado no log
            sorteado.ja_sorteado = True  # Define o atributo ja_sorteado como verdadeiro
            sorteado.save()  # Salva a alteração

        logger.info('Fim da turma.')  # Mostra o final do sorteio da turma atual no log

        """
        Por ser um loop, o processo se repetirá para todas as turmas 
        """

    logger.info("Fim do sorteio")  # Mostra o fim do sorteio no log

@shared_task
def preparar_emails_sorteio(emails_inscritos, content):
    tamanho_lote = 100
    subject = 'Incubadora de Robótica'
    tarefas = [
        enviar_emails.s(
            emails=emails_inscritos[i:i+tamanho_lote if i+tamanho_lote < len(emails_inscritos) else -1],
            content=content,
            subject=subject,
            has_html=True
        )
        for i in range(0, len(emails_inscritos), tamanho_lote)
    ]
    group(tarefas).apply_async()


'''
    PREENCHIMENTO E ENVIO DE LOGS
'''
def log_action(action, instance):
    logger = getLogger('register')
    user = get_current_user()
    username = user.username if user else "Desconhecido"
    logger.info(f'{action}: {instance._meta.model_name} | Usuário: {username} | ID: {instance.id} - {instance.__str__()}')

@shared_task
def preparar_log(email):
    log_file_path = os.path.join(settings.BASE_DIR, 'logs/register.log')

    subject = 'Log de registro'
    content = 'Segue em anexo o log de registro das unidades Centro e Inoã.'

    enviar_emails.delay(
        emails=[email],
        content=content,
        subject=subject,
        file_path=log_file_path
    )

'''
    GERAÇÃO E ENVIO DE PLANILHAS
'''
@shared_task
def preparar_planilhas(email):
    planilhas_dir = os.path.join(settings.MEDIA_ROOT, 'planilhas')

    if not os.path.exists(planilhas_dir):
        print('criando planilhas_dir')
        os.makedirs(planilhas_dir)
    print('planilhas_dir existe')

    filenames = os.listdir(planilhas_dir)

    if not filenames:
        print('planilhas vazio')
        gerar_planilhas()
        gerar_planilha_coord()
        filenames = os.listdir(planilhas_dir)
        if not filenames:
            print('não foi possível gerar planilhas')
        print('planilhas criadas')
    print('planilhas não vazio')

    zip_subdir = 'planilhas'
    zip_filename = os.path.join(settings.MEDIA_ROOT, f'{zip_subdir}/{zip_subdir}.zip')

    with zipfile.ZipFile(zip_filename, 'w') as zip_file:
        for filename in filenames:
            file_path = os.path.join(planilhas_dir, filename)
            if os.path.exists(file_path):
                with open(file_path, 'rb') as file:
                    zip_path = os.path.join(zip_subdir, filename)
                    zip_file.writestr(zip_path, file.read())
            else:
                print(f'{filename} não encontrado')

    subject = 'Planilhas pedagógicas'
    content = 'Segue em anexo um arquivo zipado com as planilhas de presença.'

    enviar_emails.delay(
        emails=[email],
        content=content,
        subject=subject,
        file_path=zip_path,
    )


def ajustar_colunas(ws):
    """
    Ajusta a largura das colunas com base no conteúdo.
    """
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Obtém a letra da coluna
        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

@shared_task
def criar_folha_presenca(wb, turmas, aulas_inicio, aulas_fim):
    """
    Cria a folha de presença no workbook.
    """
    ws_presenca = wb.create_sheet(title='Presença')

    # Definir estilos para a linha escura
    dark_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")

    start_row = 1
    sabados = []
    domingos = []

    # Cabeçalhos Gerais
    headers = [
        ('Presente', 'P', '6aa84f'),
        ('Faltoso', 'F', 'cc0000'),
        ('Faltas Justificadas', 'J', 'ff9900'),  # Nova legenda 'J'
        ('Fim de Semana', 'S/D', '3d85c6'),
        ('Feriado', 'H', '8e7cc3')
    ]

    current_column = 3  # Começando da coluna 3
    for header, short, color in headers:
        ws_presenca.merge_cells(start_row=start_row, start_column=current_column, end_row=start_row,
                                end_column=current_column + 1)
        ws_presenca.cell(row=start_row, column=current_column).value = header
        ws_presenca.cell(row=start_row, column=current_column).font = Font(bold=True)
        ws_presenca.cell(row=start_row, column=current_column).alignment = Alignment(horizontal="center")

        ws_presenca.cell(row=start_row, column=current_column + 2).value = short
        ws_presenca.cell(row=start_row, column=current_column + 2).font = Font(bold=True)
        ws_presenca.cell(row=start_row, column=current_column + 2).alignment = Alignment(horizontal="center")
        ws_presenca.cell(row=start_row, column=current_column + 2).fill = PatternFill(start_color=color,
                                                                                      end_color=color,
                                                                                      fill_type="solid")

        current_column += 4  # Avança para a próxima seção

    # Títulos de Coluna
    ws_presenca.cell(row=start_row + 1, column=1).value = 'ID'
    ws_presenca.cell(row=start_row + 1, column=2).value = 'Nome'

    total_aulas = (aulas_fim - aulas_inicio).days + 1  # Inclusivo
    data_aulas = [aulas_inicio + timedelta(days=i) for i in range(total_aulas)]

    dia_coluna_map = {}

    for i, dia_aula in enumerate(data_aulas, start=3):
        ws_presenca.cell(row=start_row + 1, column=i).value = dia_aula.strftime('%d/%m')
        ws_presenca.cell(row=start_row + 1, column=i).font = Font(bold=True)
        ws_presenca.cell(row=start_row + 1, column=i).alignment = Alignment(horizontal="center")

        if dia_aula.weekday() == 5:  # Sábado
            sabados.append(i)
        elif dia_aula.weekday() == 6:  # Domingo
            domingos.append(i)

    # Formatação de Cabeçalho
    for col in range(1, 3 + total_aulas):
        cell = ws_presenca.cell(row=start_row + 1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Regras de Formatação Condicional
    range_str = f"C3:{get_column_letter(ws_presenca.max_column)}1000"  # Ajuste o número de linhas conforme necessário

    # Definir as cores para as legendas
    red_fill = PatternFill(start_color="cc0000", end_color="cc0000", fill_type="solid")
    green_fill = PatternFill(start_color="6aa84f", end_color="6aa84f", fill_type="solid")
    orange_fill = PatternFill(start_color="ff9900", end_color="ff9900", fill_type="solid")
    blue_fill = PatternFill(start_color="3d85c6", end_color="3d85c6", fill_type="solid")
    purple_fill = PatternFill(start_color="8e7cc3", end_color="8e7cc3", fill_type="solid")

    bold_font = Font(bold=True)

    # Regras para 'F', 'P', 'J'
    ws_presenca.conditional_formatting.add(range_str,
                                           CellIsRule(operator='equal', formula=['"F"'], fill=red_fill, font=bold_font))
    ws_presenca.conditional_formatting.add(range_str,
                                           CellIsRule(operator='equal', formula=['"P"'], fill=green_fill, font=bold_font))
    ws_presenca.conditional_formatting.add(range_str,
                                           CellIsRule(operator='equal', formula=['"J"'], fill=orange_fill, font=bold_font))
    ws_presenca.conditional_formatting.add(range_str,
                                           CellIsRule(operator='equal', formula=['"S"'], fill=blue_fill, font=bold_font))
    ws_presenca.conditional_formatting.add(range_str,
                                           CellIsRule(operator='equal', formula=['"D"'], fill=blue_fill, font=bold_font))
    ws_presenca.conditional_formatting.add(range_str,
                                           CellIsRule(operator='equal', formula=['"H"'], fill=purple_fill, font=bold_font))

    # Adicionar bordas grossas para a linha escura
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Preenchimento das Turmas
    current_row = start_row + 2
    for turma in turmas:
        # Título da Turma
        header = f'{turma.dias} {turma.horario()}'
        ws_presenca.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        ws_presenca.cell(row=current_row, column=1).value = header
        ws_presenca.cell(row=current_row, column=1).font = bold_font
        ws_presenca.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
        current_row += 1

        # Alunos da Turma
        alunos = Aluno.objects.filter(id_turma=turma).annotate(
            nome_ordenacao=Coalesce('nome_social', 'nome')).order_by('nome_ordenacao')

        for aluno in alunos:
            ws_presenca.cell(row=current_row, column=1).value = aluno.pk
            ws_presenca.cell(row=current_row, column=2).value = aluno.nome_social if aluno.nome_social else aluno.nome
            ws_presenca.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
            ws_presenca.cell(row=current_row, column=2).alignment = Alignment(horizontal="left")

            # Inicializa as células de presença
            for col in range(3, 3 + total_aulas):
                ws_presenca.cell(row=current_row, column=col).value = ""
                ws_presenca.cell(row=current_row, column=col).alignment = Alignment(horizontal="center")

            current_row += 1

        # Inserir uma linha escura após a turma
        ws_presenca.append([])  # Adiciona uma linha vazia
        for col in range(1, 3 + total_aulas):
            cell = ws_presenca.cell(row=current_row, column=col)
            cell.fill = dark_fill
            cell.border = thin_border
        current_row += 1

    ajustar_colunas(ws_presenca)
    return ws_presenca

@shared_task
def criar_folha_alunos(wb, turmas):
    """
    Cria a folha de alunos no workbook.
    """
    ws_alunos = wb.create_sheet(title='Alunos')
    headers = ['ID', 'Nome', 'CPF', 'Celular', 'Rua', 'Número', 'Bairro', 'Cidade', 'UF', 'É PCD?']
    for col_num, header in enumerate(headers, start=1):
        ws_alunos.cell(row=1, column=col_num).value = header
        ws_alunos.cell(row=1, column=col_num).font = Font(bold=True)
        ws_alunos.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")

    alunos = Aluno.objects.filter(id_turma__in=turmas).annotate(
        nome_ordenacao=Coalesce('nome_social', 'nome')).order_by('nome_ordenacao')

    for row_num, aluno in enumerate(alunos, start=2):
        ws_alunos.cell(row=row_num, column=1).value = aluno.pk
        ws_alunos.cell(row=row_num, column=2).value = aluno.nome_social if aluno.nome_social else aluno.nome
        ws_alunos.cell(row=row_num, column=3).value = aluno.cpf_formatado()
        ws_alunos.cell(row=row_num, column=4).value = aluno.celular
        ws_alunos.cell(row=row_num, column=5).value = aluno.rua
        ws_alunos.cell(row=row_num, column=6).value = aluno.numero
        ws_alunos.cell(row=row_num, column=7).value = aluno.bairro
        ws_alunos.cell(row=row_num, column=8).value = aluno.cidade
        ws_alunos.cell(row=row_num, column=9).value = aluno.uf
        ws_alunos.cell(row=row_num, column=10).value = 'PCD' if aluno.pcd else ''

        # Alinhamento das células
        for col in range(1, 11):
            ws_alunos.cell(row=row_num, column=col).alignment = Alignment(horizontal="left")

    ajustar_colunas(ws_alunos)
    return ws_alunos

@shared_task
def gerar_planilhas():
    """
    Gera planilhas Excel para cada professor e curso ministrado.
    Cada planilha contém duas folhas: Presença e Alunos.
    """
    try:
        controle = Controle.objects.first()
        if not controle:
            print("Nenhum controle encontrado.")
            return

        aulas_inicio = controle.aulas_inicio
        aulas_fim = controle.aulas_fim

        curso_arq = {
            'Informática para Iniciantes e Melhor Idade': 'info_melhor_idade',
            'Informática Básica': 'info_basica',
            'Excel': 'excel',
            'Montagem e Manutenção de Computadores': 'montagem',
            'Robótica e Automação: Módulo 01': 'robotica_01',
            'Robótica e Automação: Módulo 02': 'robotica_02',
            'Robótica e Automação: Módulo 03': 'robotica_03',
            'Design e Modelagem 3D: Módulo 01': 'design_01',
            'Design e Modelagem 3D: Módulo 02': 'design_02',
            'Gestão de Pessoas': 'gestao_pessoas',
            'Educação Financeira': 'educ_finan',
            'Marketing Digital': 'mark_digital',
            'Marketing Empreendedor': 'mark_emp'
        }

        turmas = Turma.objects.all()
        professores = set(turma.professor for turma in turmas)  # Elimina duplicatas

        for professor in professores:
            # Filtra turmas do professor
            turmas_professor = turmas.filter(professor=professor)
            cursos = set(turma.curso for turma in turmas_professor)  # Elimina duplicatas

            for curso in cursos:
                if curso not in curso_arq:
                    print(f"Curso '{curso}' não mapeado em curso_arq. Pulando...")
                    continue  # Pula cursos que não estão mapeados

                # Filtra turmas específicas do curso e professor
                turmas_aux = turmas_professor.filter(curso=curso)

                if not turmas_aux.exists():
                    continue  # Pula se não houver turmas

                wb = Workbook()
                # Remove a folha padrão criada pelo Workbook
                default_sheet = wb.active
                wb.remove(default_sheet)

                # Cria as folhas necessárias
                criar_folha_presenca.delay(wb, turmas_aux, aulas_inicio, aulas_fim)
                criar_folha_alunos.delay(wb, turmas_aux)

                # Definir caminho e nome do arquivo
                nome_arquivo = f"presenca_{professor}_{curso_arq[curso]}.xlsx"
                caminho_arquivo = os.path.join('media', 'planilhas', nome_arquivo)

                # Garantir que o diretório exista
                os.makedirs(os.path.dirname(caminho_arquivo), exist_ok=True)

                # Salvar o workbook
                wb.save(caminho_arquivo)
                print(f"Planilha salva: {caminho_arquivo}")

    except Exception as e:
        print(f"Ocorreu um erro durante a geração das planilhas: {e}")

@shared_task
def criar_folha_geral(wb: Workbook):
    ws_geral = wb.create_sheet(title='Geral')
    headers = ['ID', 'Nome', 'CPF', 'Celular', 'Rua', 'Número', 'Bairro', 'Cidade', 'UF', 'É PCD?', 'OBS', 'Curso', 'Dias', 'Horario']
    for col_num, header in enumerate(headers, start=1):
        ws_geral.cell(row=1, column=col_num).value = header
        ws_geral.cell(row=1, column=col_num).font = Font(bold=True)
        ws_geral.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")

    alunos = Aluno.objects.all()

    for row_num, aluno in enumerate(alunos, start=2):
        turma = aluno.id_turma

        ws_geral.cell(row=row_num, column=1).value = aluno.pk
        ws_geral.cell(row=row_num, column=2).value = aluno.nome_social if aluno.nome_social else aluno.nome
        ws_geral.cell(row=row_num, column=3).value = aluno.cpf_formatado()
        ws_geral.cell(row=row_num, column=4).value = aluno.celular
        ws_geral.cell(row=row_num, column=5).value = aluno.rua
        ws_geral.cell(row=row_num, column=6).value = aluno.numero
        ws_geral.cell(row=row_num, column=7).value = aluno.bairro
        ws_geral.cell(row=row_num, column=8).value = aluno.cidade
        ws_geral.cell(row=row_num, column=9).value = aluno.uf
        ws_geral.cell(row=row_num, column=10).value = 'PCD' if aluno.pcd else ''
        ws_geral.cell(row=row_num, column=11).value = aluno.observacoes if aluno.observacoes is not None else ''
        ws_geral.cell(row=row_num, column=12).value = turma.curso
        ws_geral.cell(row=row_num, column=13).value = turma.dias
        ws_geral.cell(row=row_num, column=14).value = turma.horario()

        for col in range(1, 15):
            ws_geral.cell(row=row_num, column=col).alignment = Alignment(horizontal="left")

    ajustar_colunas(ws_geral)
    return ws_geral

@shared_task
def criar_folha_curso(wb: Workbook, curso: str, turmas: QuerySet[Turma]):
    dark_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")

    ws_curso = wb.create_sheet(title=f'{curso}')
    headers = ['ID', 'Nome', 'CPF', 'Celular', 'Rua', 'Número', 'Bairro', 'Cidade', 'UF', 'É PCD?', 'OBS']

    init_row = 1
    for turma in turmas:
        last_row = init_row

        # Separação de turmas
        ws_curso.cell(row=init_row, column=1).fill = dark_fill
        ws_curso.cell(row=init_row, column=2).value = f'{turma.dias} - {turma.horario()}'
        for col in range(3, 12):
            ws_curso.cell(row=init_row, column=col).fill = dark_fill

        # Headers
        for col_num, header in enumerate(headers, start=1):
            ws_curso.cell(row=init_row + 1, column=col_num).value = header
            ws_curso.cell(row=init_row + 1, column=col_num).font = Font(bold=True)
            ws_curso.cell(row=init_row + 1, column=col_num).alignment = Alignment(horizontal="center")

        alunos = (Aluno.objects.filter(id_turma=turma)
                  .annotate(nome_ordenacao=Coalesce('nome_social', 'nome'))
                  .order_by('nome_ordenacao'))

        last_row += 2
        for row_num, aluno in enumerate(alunos, start=last_row):
            ws_curso.cell(row=row_num, column=1).value = aluno.pk
            ws_curso.cell(row=row_num, column=2).value = aluno.nome_social if aluno.nome_social else aluno.nome
            ws_curso.cell(row=row_num, column=3).value = aluno.cpf_formatado()
            ws_curso.cell(row=row_num, column=4).value = aluno.celular
            ws_curso.cell(row=row_num, column=5).value = aluno.rua
            ws_curso.cell(row=row_num, column=6).value = aluno.numero
            ws_curso.cell(row=row_num, column=7).value = aluno.bairro
            ws_curso.cell(row=row_num, column=8).value = aluno.cidade
            ws_curso.cell(row=row_num, column=9).value = aluno.uf
            ws_curso.cell(row=row_num, column=10).value = 'PCD' if aluno.pcd else ''
            ws_curso.cell(row=row_num, column=11).value = aluno.observacoes if aluno.observacoes is not None else ''

            for col in range(1, 12):
                ws_curso.cell(row=row_num, column=col).alignment = Alignment(horizontal="left")
            last_row += 1

        for col_num in range(1, 12):
            ws_curso.cell(row=last_row, column=col_num).fill = dark_fill

        init_row = last_row + 1

    print(ws_curso)
    ajustar_colunas(ws_curso)
    return ws_curso


def gerar_planilha_coord():
    curso_title = {
        'Informática para Iniciantes e Melhor Idade': 'Info Iniciante',
        'Informática Básica': 'Info Básica',
        'Excel': 'Excel',
        'Montagem e Manutenção de Computadores': 'Montagem',
        'Robótica e Automação: Módulo 01': 'Robótica 1',
        'Robótica e Automação: Módulo 02': 'Robótica 2',
        'Robótica e Automação: Módulo 03': 'Robótica 3',
        'Design e Modelagem 3D: Módulo 01': 'Design 1',
        'Design e Modelagem 3D: Módulo 02': 'Design 2',
        'Gestão de Pessoas': 'Gestão',
        'Educação Financeira': 'Educação Financeira',
        'Marketing Digital': 'M. Digital',
        'Marketing Empreendedor': 'M. Empreendedor'
    }

    try:
        wb = Workbook()

        default_sheet = wb.active
        wb.remove(default_sheet)

        criar_folha_geral.delay(wb)

        turmas = Turma.objects.all()
        cursos = set(turma.curso for turma in turmas)

        for curso in cursos:
            turmas_aux = turmas.filter(curso=curso)

            if not turmas_aux.exists():
                continue

            criar_folha_curso.delay(wb, curso_title[curso], turmas_aux)

        nome_arquivo = f'planilha_geral_2024_2.xlsx'
        caminho_arquivo = os.path.join('media', 'planilhas', nome_arquivo)

        os.makedirs(os.path.dirname(caminho_arquivo), exist_ok=True)

        wb.save(caminho_arquivo)
        print(f'Planilha salva: {caminho_arquivo}')

    except Exception as e:
        print(f'Ocorreu um erro durante a geração da planilha: {e}')


'''
    AGENDAMENTO DE TAREFAS
'''
@shared_task
def agendar_task(task_name, task_path, eta, parametro=None):
    schedule, _ = IntervalSchedule.objects.get_or_create(
        every=1,
        period=IntervalSchedule.MINUTES,
    )

    args_json = json.dumps([parametro]) if parametro is not None else "[]"

    task, created = PeriodicTask.objects.get_or_create(
        name=task_name,
        defaults={
            'interval': schedule,
            'task': task_path,
            'args': args_json,
            'start_time': eta,
            'one_off': True,
        }
    )

    if not created:
        task.start_time = eta
        task.args = args_json
        task.save()
        print(f'Tarefa {task_name} atualizada para {eta}.')
    else:
        print(f'Tarefa {task_name} agendada para {eta}.')
