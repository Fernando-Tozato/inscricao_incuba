from datetime import timedelta

from django.db.models import Q, QuerySet
from django.db.models.functions import Coalesce
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from database.models import *


class Presenca:
    ws: Worksheet
    professor: str
    curso: str

    def __init__(self, ws: Worksheet, professor: str, curso: str):
        self.ws = ws
        self.professor = professor
        self.curso = curso

    def generate(self) -> Worksheet:
        turmas: QuerySet[Turma] = Turma.objects.filter(Q(professor=self.professor) & Q(curso=self.curso))

        start_row: int = 1
        for turma in turmas:
            header = f'{turma.dias} {turma.horario()}'
            self.ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
            self.ws.cell(row=start_row, column=1).value = header
            self.ws.cell(row=start_row, column=1).font = Font(bold=True)
            self.ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center')

            self.gerar_legenda(start_row, 3, 'Presente', 'P', '6aa84f')  # Presente
            self.gerar_legenda(start_row, 7, 'Faltoso', 'F', 'cc0000')  # Faltoso
            self.gerar_legenda(start_row, 11, 'Fim de Semana', 'S/D', '3d85c6')  # Fim de Semana
            self.gerar_legenda(start_row, 15, 'Feriado', 'H', 'f1c232')  # Feriado

            self.gerar_col_header(start_row)

            self.preencher_alunos(start_row, turma)

            end_row: int = self.ws.max_row

            range_str = f'C{start_row + 2}:{get_column_letter(self.ws.max_column)}{end_row + start_row + 2}'
            self.formatar_presenca(range_str)

            max_row = end_row + start_row + 3

            self.formatar_espacamento(max_row)

            start_row = max_row + 1

        return self.ws

    def gerar_legenda(self, start_row, start_col, name, char, color):
        self.ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)

        start_cell = self.ws.cell(row=start_row, column=start_col)
        end_cell = self.ws.cell(row=start_row, column=start_col + 2)

        start_cell.value = name
        start_cell.font = Font(bold=True)
        start_cell.alignment = Alignment(horizontal="center")
        end_cell.value = char
        end_cell.font = Font(bold=True)
        end_cell.alignment = Alignment(horizontal="center")
        end_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    def gerar_col_header(self, start_row):
        aulas_inicio = Controle.objects.first().aulas_inicio
        aulas_fim = Controle.objects.first().aulas_fim

        self.ws.cell(row=start_row + 1, column=1).value = 'ID'
        self.ws.cell(row=start_row + 1, column=2).value = 'Nome'

        for col in range(3, (aulas_fim - aulas_inicio).days + 4):
            dia_aula = aulas_inicio + timedelta(days=col - 3)

            cell = self.ws.cell(row=start_row + 1, column=col)

            cell.value = dia_aula.strftime('%d/%m')
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    def preencher_alunos(self, start_row, turma):
        alunos = (Aluno.objects.filter(id_turma=turma)
                  .annotate(nome_ordenacao=Coalesce('nome_social', 'nome'))
                  .order_by('nome_ordenacao'))

        idx = 0

        for idx, aluno in enumerate(alunos):
            self.ws.cell(row=idx + start_row + 2, column=1).value = aluno.pk
            self.ws.cell(row=idx + start_row + 2,
                         column=2).value = aluno.nome_social if aluno.nome_social else aluno.nome

            self.preencher_cells(start_row)

        return idx

    def formatar_presenca(self, range_str):
        red_fill = PatternFill(start_color="cc0000", end_color="cc0000", fill_type="solid")
        green_fill = PatternFill(start_color="6aa84f", end_color="6aa84f", fill_type="solid")
        yellow_fill = PatternFill(start_color="f1c232", end_color="f1c232", fill_type="solid")

        self.ws.conditional_formatting.add(range_str, CellIsRule(operator='equal', formula=['"F"'], fill=red_fill))
        self.ws.conditional_formatting.add(range_str, CellIsRule(operator='equal', formula=['"P"'], fill=green_fill))
        self.ws.conditional_formatting.add(range_str, CellIsRule(operator='equal', formula=['"H"'], fill=yellow_fill))

    def formatar_espacamento(self, max_row):
        for col in range(1, self.ws.max_column + 1):
            self.ws.cell(row=max_row, column=col).fill = PatternFill(start_color="000000", end_color="000000",
                                                                     fill_type="solid")

    def preencher_cells(self, row):
        aulas_inicio = Controle.objects.first().aulas_inicio
        aulas_fim = Controle.objects.first().aulas_fim

        for col in range(3, (aulas_fim - aulas_inicio).days + 4):
            dia_aula = aulas_inicio + timedelta(days=col - 3)
            cell = self.ws.cell(row=row, column=col)
            if dia_aula.weekday() == 5:
                cell.value = 'S'
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color='3d85c6', end_color='3d85c6', fill_type='solid')
            elif dia_aula.weekday() == 6:
                cell.value = 'D'
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color='3d85c6', end_color='3d85c6', fill_type='solid')
