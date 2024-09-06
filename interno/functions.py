import threading
from datetime import timedelta

from django.db.models import Q
from django.db.models.functions import Coalesce
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from database.models import *


def ja_matriculado(cpf):
    return Aluno.objects.filter(cpf=cpf).exists()


def formatar_presenca(ws, range_str):
    red_fill = PatternFill(start_color="cc0000", end_color="cc0000", fill_type="solid")
    green_fill = PatternFill(start_color="6aa84f", end_color="6aa84f", fill_type="solid")
    yellow_fill = PatternFill(start_color="f1c232", end_color="f1c232", fill_type="solid")

    ws.conditional_formatting.add(range_str, CellIsRule(operator='equal', formula=['"F"'], fill=red_fill))
    ws.conditional_formatting.add(range_str, CellIsRule(operator='equal', formula=['"P"'], fill=green_fill))
    ws.conditional_formatting.add(range_str, CellIsRule(operator='equal', formula=['"H"'], fill=yellow_fill))


def formatar_espacamento(ws, max_row):
    for col in range(1, ws.max_column + 1):
        ws.cell(row=max_row, column=col).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")


def ajustar_colunas(ws):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Obtém a letra da coluna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


def gerar_legenda(ws, start_row, start_col, name, char, color):
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)

    start_cell = ws.cell(row=start_row, column=start_col)
    end_cell = ws.cell(row=start_row, column=start_col + 2)

    start_cell.value = name
    start_cell.font = Font(bold=True)
    start_cell.alignment = Alignment(horizontal="center")
    end_cell.value = char
    end_cell.font = Font(bold=True)
    end_cell.alignment = Alignment(horizontal="center")
    end_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def gerar_col_header(ws, start_row, ):
    aulas_inicio = Controle.objects.first().aulas_inicio
    aulas_fim = Controle.objects.first().aulas_fim

    ws.cell(row=start_row + 1, column=1).value = 'ID'
    ws.cell(row=start_row + 1, column=2).value = 'Nome'

    for col in range(3, (aulas_fim - aulas_inicio).days + 4):
        dia_aula = aulas_inicio + timedelta(days=col - 3)

        cell = ws.cell(row=start_row + 1, column=col)

        cell.value = dia_aula.strftime('%d/%m')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def preencher_cells(ws, row):
    aulas_inicio = Controle.objects.first().aulas_inicio
    aulas_fim = Controle.objects.first().aulas_fim

    for col in range(3, (aulas_fim - aulas_inicio).days + 4):
        dia_aula = aulas_inicio + timedelta(days=col - 3)
        cell = ws.cell(row=row, column=col)
        if dia_aula.weekday() == 5:
            cell.value = 'S'
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color='3d85c6', end_color='3d85c6', fill_type='solid')
        elif dia_aula.weekday() == 6:
            cell.value = 'D'
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color='3d85c6', end_color='3d85c6', fill_type='solid')


def preencher_alunos(ws, start_row, turma):
    alunos = (Aluno.objects.filter(id_turma=turma)
              .annotate(nome_ordenacao=Coalesce('nome_social', 'nome'))
              .order_by('nome_ordenacao'))

    idx = 0

    for idx, aluno in enumerate(alunos):
        ws.cell(row=idx + start_row + 2, column=1).value = aluno.pk
        ws.cell(row=idx + start_row + 2, column=2).value = aluno.nome_social if aluno.nome_social else aluno.nome

        preencher_cells(ws, start_row)

    return idx


def gerar_presenca(ws, curso, professor):
    turmas = Turma.objects.filter(Q(professor=professor) & Q(curso=curso))

    start_row = 1
    for turma in turmas:
        # Título da Turma
        header = f'{turma.dias} {turma.horario()}'
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
        ws.cell(row=start_row, column=1).value = header
        ws.cell(row=start_row, column=1).font = Font(bold=True)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center')

        gerar_legenda(ws, start_row, 3, 'Presente', 'P', '6aa84f')  # Presente
        gerar_legenda(ws, start_row, 7, 'Faltoso', 'F', 'cc0000')  # Faltoso
        gerar_legenda(ws, start_row, 11, 'Fim de Semana', 'S/D', '3d85c6')  # Fim de Semana
        gerar_legenda(ws, start_row, 15, 'Feriado', 'H', 'f1c232')  # Feriado

        gerar_col_header(ws, start_row)

        preencher_alunos(ws, start_row, turma)

        end_row = ws.max_row

        range_str = f'C{start_row + 2}:{get_column_letter(ws.max_column)}{end_row + start_row + 2}'
        formatar_presenca(ws, range_str)

        max_row = end_row + start_row + 3

        formatar_espacamento(ws, max_row)

        start_row = max_row + 1


def criar_headers(ws):
    headers = ['ID', 'Nome', 'CPF', 'Celular', 'Rua', 'Número', 'Bairro', 'Cidade', 'UF', 'É PCD?']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')


def gerar_alunos(ws, curso, professor):
    turmas = Turma.objects.filter(Q(professor=professor) & Q(curso=curso))
    alunos = (Aluno.objects.filter(id_turma__in=turmas)
              .annotate(nome_ordenacao=Coalesce('nome_social', 'nome'))
              .order_by('nome_ordenacao'))

    criar_headers(ws)

    for row, aluno in enumerate(alunos, start=2):
        ws.cell(row=row, column=1).value = aluno.pk
        ws.cell(row=row, column=2).value = aluno.nome_social if aluno.nome_social else aluno.nome
        ws.cell(row=row, column=3).value = aluno.cpf
        ws.cell(row=row, column=4).value = aluno.celular
        ws.cell(row=row, column=5).value = aluno.rua
        ws.cell(row=row, column=6).value = aluno.numero
        ws.cell(row=row, column=7).value = aluno.bairro
        ws.cell(row=row, column=8).value = aluno.cidade
        ws.cell(row=row, column=9).value = aluno.uf
        ws.cell(row=row, column=10).value = 'PCD' if aluno.pcd else ''


def gerar_planilhas(professor):
    curso_arq = {
        'Informática para Iniciantes e Melhor Idade': 'info_melhor_idade',
        'Informática Básica': 'info_basica',
        'Excel Intermediário': 'excel_int',
        'Excel Avançado': 'excel_ava',
        'Montagem e Manutenção de Computadores': 'montagem',
        'Robótica e Automação: Módulo 01': 'robotica_01',
        'Robótica e Automação: Módulo 02': 'robotica_02',
        'Robótica e Automação: Módulo 03': 'robotica_03',
        'Design e Modelagem 3D: Módulo 01': 'design_01',
        'Design e Modelagem 3D: Módulo 02': 'design_02',
        'Gestão de Pessoas': 'gestao',
        'Educação Financeira': 'educ_finan',
        'Marketing Digital e Empreendedor': 'mark_digital_emp',
        'Ferramentas do Marketing Digital': 'fer_mark_digital'
    }

    cursos = Turma.objects.filter(professor=professor).values_list('curso', flat=True).distinct()

    for curso in cursos:
        if not curso:
            del cursos[cursos.index(curso)]
            continue

        wb = Workbook()
        ws_presenca = wb.create_sheet(title='Presença')
        ws_alunos = wb.create_sheet(title='Alunos')

        gerar_presenca(ws_presenca, professor, curso)

        ajustar_colunas(ws_presenca)

        gerar_alunos(ws_alunos, curso, professor)

        ajustar_colunas(ws_alunos)

        filename = f'media/planilhas/presenca_{professor}_{curso_arq[curso]}.xlsx'
        wb.save(filename)


def criar_threads():
    professores = Turma.objects.values_list('professor', flat=True).distinct()

    threads = []

    for professor in professores:
        thread = threading.Thread(target=gerar_planilhas, args=(professor,))
        threads.append(thread)

    for thread in threads:
        thread.start()

    for thread in threads:
        thread.join()
