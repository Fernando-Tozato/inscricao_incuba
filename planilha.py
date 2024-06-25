import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from django.db.models import Q
from datetime import datetime, timedelta
from your_app.models import Turma, Aluno

def gerar_planilha_presenca(professor, data_inicio, data_fim):
    # Criar um novo workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remover a folha padrão

    # Converter strings para objetos de data
    data_inicio = datetime.strptime(data_inicio, "%Y-%m-%d")
    data_fim = datetime.strptime(data_fim, "%Y-%m-%d")

    # Buscar todas as turmas do professor no intervalo de datas
    turmas = Turma.objects.filter(professor=professor)
    dias_semana = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]

    for turma in turmas:
        # Criar uma nova folha para cada turma
        sheet_name = f"{turma.curso} ({turma.dias})"
        ws = wb.create_sheet(title=sheet_name[:31])  # Limitar a 31 caracteres

        # Título da planilha
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Presença - {turma.curso}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")

        # Cabeçalho
        ws.append(["ID", "Nome", "Contato", "Endereço"] + [(data_inicio + timedelta(days=i)).strftime("%d/%m") for i in range((data_fim - data_inicio).days + 1)])

        # Estilo do cabeçalho
        for col in range(1, ws.max_column + 1):
            ws.cell(row=2, column=col).font = Font(bold=True)
            ws.cell(row=2, column=col).alignment = Alignment(horizontal="center")

        # Buscar alunos da turma
        alunos = Aluno.objects.filter(turma=turma)

        # Preencher dados dos alunos
        for aluno in alunos:
            ws.append([
                aluno.id,
                aluno.nome,
                aluno.contato,
                aluno.endereco
            ] + [""] * ((data_fim - data_inicio).days + 1))

        # Ajustar largura das colunas
        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15

    # Salvar o workbook
    filename = f"presenca_{professor}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

# Exemplo de uso:
# gerar_planilha_presenca('Cláudia', '2023-06-01', '2023-06-30')
